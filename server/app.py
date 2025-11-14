# server/app.py
import os
import json
import io
import tempfile
import logging
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pymongo
from gridfs import GridFS
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from graph_upload_server import analyze_chart
from extract_toc import find_toc_page_range, extract_toc_entries, build_topic_map
from spreadsheet_analysis import analyze_spreadsheet_auto_merge

# --- CONFIG ---
MONGO_URI    = "mongodb://localhost:27017/"
DB_NAME      = "pdf_bot"
PDF_BUCKET   = "pdfs"       # GridFS bucket name
MAPPING_COLL = "mappings"   # Stores { topic_map: {...} }
INDEX_COLL   = "index"      # Stores { filename, mapping_id }
# ----------------

# set up logging
logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Mongo connection
client = pymongo.MongoClient(MONGO_URI)
db     = client[DB_NAME]
fs     = GridFS(db, collection=PDF_BUCKET)
maps   = db[MAPPING_COLL]
idx    = db[INDEX_COLL]

@app.route("/upload", methods=["POST"])
def upload_file():
    log.info("Received /upload request")
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files["file"]
    filename = file.filename
    if not filename or not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Invalid file", "ok": False}), 400

    try:
        existing = fs.find_one({"filename": filename})
        if existing:
            fs.delete(existing._id)

        pdf_bytes = file.read()
        pdf_id = fs.put(pdf_bytes, filename=filename)

        fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
        os.close(fd)
        with open(tmp_path, "wb") as tmp:
            tmp.write(pdf_bytes)

        toc_start, toc_end = find_toc_page_range(tmp_path)

        if toc_start:
            entries = extract_toc_entries(tmp_path, toc_start, toc_end)
            topic_map = build_topic_map(entries)
            log.info("Extracted %d ToC entries", len(topic_map))
        else:
            topic_map = {}
            log.warning("No TOC detected in %s. Saving empty topic_map.", filename)

        os.remove(tmp_path)

        mapping_doc = {"topic_map": topic_map}
        mapping_id = maps.insert_one(mapping_doc).inserted_id

        idx.replace_one(
            {"filename": filename},
            {"filename": filename, "mapping_id": mapping_id},
            upsert=True
        )

        return jsonify({
            "message": "Upload successful",
            "filename": filename,
            "mapping_id": str(mapping_id),
            "ok": True,
            "toc_found": bool(toc_start)
        }), 200

    except Exception as e:
        log.exception("Error in upload")
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        return jsonify({"error": f"Server error: {e}"}), 500

@app.route("/list_pdfs", methods=["GET"])
def list_pdfs():
    docs = idx.find({}, {"filename": 1, "_id": 0})
    filenames = [doc["filename"] for doc in docs]
    return jsonify({"pdfs": filenames, "ok": True})

@app.route("/upload_graph", methods=["POST"])
def upload_graph():
    log.info("Received /upload_graph request")
    if "file" not in request.files:
        return jsonify({"error": "No file part", "ok": False}), 400

    file = request.files["file"]
    filename = secure_filename(file.filename)
    if not filename.lower().endswith((".png", ".jpg", ".jpeg")):
        return jsonify({"error": "Invalid file type. Only PNG/JPG allowed.", "ok": False}), 400

    try:
        fd, tmp_path = tempfile.mkstemp(suffix=os.path.splitext(filename)[1])
        os.close(fd)
        with open(tmp_path, "wb") as tmp_file:
            tmp_file.write(file.read())

        result = analyze_chart(tmp_path)
        raw_content = result["choices"][0]["message"]["content"]
        log.info("Azure raw response:\n%s", raw_content)

        import re
        json_str = re.sub(r"```json|```", "", raw_content).strip()

        data = json.loads(json_str)
        os.remove(tmp_path)

        raw_pts = data.get("data_points") or data.get("dataPoints") or []
        if isinstance(raw_pts, list):
            flat_points = raw_pts
        elif isinstance(raw_pts, dict):
            flat_points = [{"label": k, "value": v} for k, v in raw_pts.items()]
        else:
            flat_points = []

        return jsonify({
            "ok": True,
            "raw": data,  # full original parsed content
            "title": data.get("title"),
            "x_axis_label": data.get("x_axis_label") or data.get("axes", {}).get("x"),
            "y_axis_label": data.get("y_axis_label") or data.get("axes", {}).get("y"),
            "data": data.get("data", []),
            "dataPoints": flat_points
        }), 200


    except Exception as e:
        log.exception("Error in /upload_graph")
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        return jsonify({"error": f"Server error: {e}", "ok": False}), 500

@app.route("/upload_sheet", methods=["POST"])
def upload_sheet_merged():
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files["file"]
    filename = secure_filename(file.filename)
    if not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF allowed"}), 400

    pdf_bytes = file.read()
    try:
        data = analyze_spreadsheet_auto_merge(pdf_bytes, filename)
        return jsonify({"ok": True, "data": data}), 200
    except Exception as e:
        log.exception("Error in /upload_sheet/merged")
        return jsonify({"error": str(e), "ok": False}), 500

@app.route("/getexcel", methods=["POST"])
def get_excel():
    """
    Accepts JSON payload in the same format as the analyzer output,
    returns an .xlsx file for download.
    """
    data = request.get_json()
    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in data.get("sheets", []):
        ws = wb.create_sheet(title=sheet.get("name", "Sheet"))
        for row in sheet.get("rows", []):
            for cell in row.get("cells", []):
                if cell.get("enable", False):
                    ws.cell(
                        row=row.get("index"),
                        column=cell.get("index"),
                        value=cell.get("value")
                    )
    # Activate sheet
    active = data.get("activeSheet")
    if active in wb.sheetnames:
        wb.active = wb[active]
    # Save to in-memory bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    # Send as downloadable file
    return send_file(
        output,
        as_attachment=True,
        download_name="output.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(port=5001)
